"""
Export service — one-click export tools for trust data.

Phase 4 (Enhanced Features) of the TrustOffice plan.

Exports all trust data as structured JSON or ZIP archives, and creates
comprehensive backup archives stored in the vault.
"""
import io
import json
import uuid
import zipfile
from datetime import datetime, timezone
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import db


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


async def get_owned_trust(trust_id: str, user_id: str) -> Optional[dict]:
    """Fetch a trust only if owned by this user (mirrors meeting_service pattern)."""
    return await db.trusts.find_one(
        {"trust_id": trust_id, "user_id": user_id}, {"_id": 0}
    )


# ==================== COLLECTION MAP ====================

# Maps export category → (collection_name, extra_filter_fields)
# All collections are scoped by trust_id + user_id.
_EXPORT_COLLECTIONS = {
    "trust_details": ("trusts", None),
    "entities": ("entities", None),
    "beneficiaries": ("trust_unit_certificates", None),
    "class_beneficiaries": ("class_beneficiaries", None),
    "unit_transfers": ("trust_unit_transfers", None),
    "units_settings": ("trust_units_settings", None),
    "minutes": ("minutes_records", None),
    "meetings": ("meetings", None),
    "deadlines": ("deadlines", None),
    "tasks": ("governance_tasks", None),
    "distributions": ("distribution_records", None),
    "compensation": ("compensation_payments", None),
    "expenses": ("expenses", None),
    "schedule_a": ("schedule_a_items", None),
    "transactions": ("transactions", None),
    "health_snapshots": ("health_score_snapshots", None),
    "separation_alerts": ("separation_alerts", None),
    "audit_log": ("audit_log", None),
    "communications": ("communications", None),
    "calendar_events": ("calendar_events", None),
    "notifications": ("notifications", None),
}


async def _export_collection(
    collection_name: str, trust_id: str, user_id: str, exclude_file_content: bool = True
) -> list:
    """Export all documents from a collection for a trust."""
    projection = {"_id": 0}
    if exclude_file_content and collection_name == "vault_documents":
        projection["file_content"] = 0

    return await db[collection_name].find(
        {"trust_id": trust_id, "user_id": user_id},
        projection,
    ).to_list(10000)


async def export_trust_data(
    trust_id: str, user_id: str, format: str = "json"
) -> dict:
    """Export ALL trust data as structured JSON or ZIP.

    For JSON format: returns {format, exported_at, trust_id, trust_name, data: {...}}.
    For ZIP format: returns {format, exported_at, trust_id, trust_name, zip_bytes, filename}.

    Raises ValueError if trust not found.
    """
    trust = await get_owned_trust(trust_id, user_id)
    if not trust:
        raise ValueError("Trust not found")

    trust_name = trust.get("name", "Unnamed Trust")
    exported_at = _now()

    # Gather all data
    data = {}
    for category, (collection_name, _) in _EXPORT_COLLECTIONS.items():
        try:
            data[category] = await _export_collection(
                collection_name, trust_id, user_id
            )
        except Exception:
            data[category] = []

    # Vault documents metadata (no file contents)
    try:
        data["vault_documents"] = await _export_collection(
            "vault_documents", trust_id, user_id, exclude_file_content=True
        )
    except Exception:
        data["vault_documents"] = []

    if format == "zip":
        zip_bytes = _build_zip(trust_name, exported_at, data)
        return {
            "format": "zip",
            "exported_at": exported_at,
            "trust_id": trust_id,
            "trust_name": trust_name,
            "zip_bytes": zip_bytes,
            "filename": f"trust_export_{trust_id}_{exported_at[:10]}.zip",
            "categories": {k: len(v) for k, v in data.items()},
        }

    return {
        "format": "json",
        "exported_at": exported_at,
        "trust_id": trust_id,
        "trust_name": trust_name,
        "data": data,
        "categories": {k: len(v) for k, v in data.items()},
    }


async def export_client_data(
    client_id: str, user_id: str, format: str = "json"
) -> dict:
    """Export all data across all trusts for a client.

    Raises ValueError if client not found.
    """
    client = await db.clients.find_one(
        {"client_id": client_id, "user_id": user_id}, {"_id": 0}
    )
    if not client:
        raise ValueError("Client not found")

    # Find all trusts for this client
    trusts = await db.trusts.find(
        {"client_id": client_id, "user_id": user_id}, {"_id": 0}
    ).to_list(100)

    exported_at = _now()
    all_trust_data = {}
    for trust in trusts:
        tid = trust["trust_id"]
        trust_data = {}
        for category, (collection_name, _) in _EXPORT_COLLECTIONS.items():
            try:
                trust_data[category] = await _export_collection(
                    collection_name, tid, user_id
                )
            except Exception:
                trust_data[category] = []
        try:
            trust_data["vault_documents"] = await _export_collection(
                "vault_documents", tid, user_id, exclude_file_content=True
            )
        except Exception:
            trust_data["vault_documents"] = []
        all_trust_data[tid] = {
            "trust_name": trust.get("name", "Unnamed Trust"),
            "data": trust_data,
        }

    client_name = client.get("name", client.get("client_name", "Unknown Client"))

    if format == "zip":
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            # Client info
            zf.writestr("client_info.json", json.dumps(client, indent=2, default=str))
            for tid, tdata in all_trust_data.items():
                safe_name = "".join(
                    c if c.isalnum() or c in " -_" else "" for c in tdata["trust_name"]
                ).strip().replace(" ", "_")
                for category, records in tdata["data"].items():
                    zf.writestr(
                        f"{safe_name}_{tid}/{category}.json",
                        json.dumps(records, indent=2, default=str),
                    )
        zip_bytes = zip_buffer.getvalue()
        return {
            "format": "zip",
            "exported_at": exported_at,
            "client_id": client_id,
            "client_name": client_name,
            "trust_count": len(trusts),
            "zip_bytes": zip_bytes,
            "filename": f"client_export_{client_id}_{exported_at[:10]}.zip",
        }

    return {
        "format": "json",
        "exported_at": exported_at,
        "client_id": client_id,
        "client_name": client_name,
        "client": client,
        "trust_count": len(trusts),
        "trusts": all_trust_data,
    }


async def create_archive_backup(trust_id: str, user_id: str) -> dict:
    """Create a comprehensive ZIP archive of the trust: all data + vault document files.

    Stores the archive in db.vault_documents with category "archive_backup".
    Returns {archive_id, doc_id, size_bytes, created_at}.
    Raises ValueError if trust not found.
    """
    trust = await get_owned_trust(trust_id, user_id)
    if not trust:
        raise ValueError("Trust not found")

    trust_name = trust.get("name", "Unnamed Trust")
    created_at = _now()
    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")

    # Export all data
    export_result = await export_trust_data(trust_id, user_id, format="json")
    data = export_result["data"]

    # Get vault documents WITH file content
    vault_docs = await db.vault_documents.find(
        {"trust_id": trust_id, "user_id": user_id, "file_content": {"$exists": True}},
    ).to_list(200)

    # Build ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Trust info
        zf.writestr(
            "trust_info.json",
            json.dumps(
                {
                    "trust_id": trust_id,
                    "trust_name": trust_name,
                    "exported_at": created_at,
                    "trust_details": data.get("trust_details", []),
                },
                indent=2,
                default=str,
            ),
        )

        # Data categories
        for category, records in data.items():
            if category == "trust_details":
                continue  # Already in trust_info.json
            zf.writestr(
                f"data/{category}.json",
                json.dumps(records, indent=2, default=str),
            )

        # Vault files
        for vdoc in vault_docs:
            file_content = vdoc.get("file_content")
            if not file_content:
                continue
            file_name = vdoc.get("file_name", f"document_{vdoc.get('doc_id', 'unknown')}")
            safe_name = "".join(
                c if c.isalnum() or c in " ._-" else "" for c in file_name
            ).strip()
            category = vdoc.get("category", "other")
            zf.writestr(f"vault/{category}/{safe_name}", file_content)

    zip_bytes = zip_buffer.getvalue()
    size_bytes = len(zip_bytes)

    # Store in vault_documents
    archive_id = _new_id("arch")
    doc_id = _new_id("doc")

    if size_bytes < 1024:
        size_display = f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        size_display = f"{size_bytes / 1024:.1f} KB"
    else:
        size_display = f"{size_bytes / (1024 * 1024):.1f} MB"

    record = {
        "doc_id": doc_id,
        "archive_id": archive_id,
        "trust_id": trust_id,
        "user_id": user_id,
        "title": f"Archive Backup — {trust_name} — {date_str}",
        "category": "archive_backup",
        "category_label": "Archive Backup",
        "date": created_at[:10],
        "description": f"Full trust archive: {len(data)} data categories, {len(vault_docs)} vault files",
        "storage_provider": "trustoffice",
        "storage_url": None,
        "storage_path": None,
        "file_name": f"archive_{trust_id}_{date_str}.zip",
        "file_size": size_display,
        "file_size_bytes": size_bytes,
        "file_content_type": "application/zip",
        "file_content": zip_bytes,
        "tags": ["archive_backup", "auto_generated"],
        "expiration_date": None,
        "needs_renewal": False,
        "created_at": created_at,
        "updated_at": created_at,
    }
    await db.vault_documents.insert_one(record)

    return {
        "archive_id": archive_id,
        "doc_id": doc_id,
        "size_bytes": size_bytes,
        "created_at": created_at,
    }


def _build_zip(trust_name: str, exported_at: str, data: dict) -> bytes:
    """Build a ZIP archive from exported data categories."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "export_info.json",
            json.dumps(
                {
                    "trust_name": trust_name,
                    "exported_at": exported_at,
                    "categories": {k: len(v) for k, v in data.items()},
                },
                indent=2,
            ),
        )
        for category, records in data.items():
            zf.writestr(
                f"{category}.json",
                json.dumps(records, indent=2, default=str),
            )
    return zip_buffer.getvalue()


# ==================== EXCEL EXPORT ====================

# Header style for Excel sheets
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row: int = 1, col_count: int = 0):
    """Apply the standard navy header style to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


async def export_health_scores_to_excel(trust_id: str, user_id: str) -> dict:
    """Export all health score snapshots for a trust to an .xlsx workbook.

    Returns {format, exported_at, trust_id, trust_name, xlsx_bytes, filename, snapshot_count}.
    Raises ValueError if trust not found.

    The workbook contains one sheet ("Health Scores") with one row per snapshot:
    calculated_at, score_value, color, base_score, risk_penalty, plus per-criterion
    columns (points / max_points / achieved) flattened from criteria_breakdown.
    """
    trust = await get_owned_trust(trust_id, user_id)
    if not trust:
        raise ValueError("Trust not found")

    trust_name = trust.get("name", "Unnamed Trust")
    exported_at = _now()

    snapshots = await db.health_score_snapshots.find(
        {"trust_id": trust_id, "user_id": user_id},
        {"_id": 0},
    ).sort("calculated_at", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is never None for a fresh Workbook
    ws.title = "Health Scores"

    # Collect the union of all criterion names (preserving first-seen order)
    criterion_names: list = []
    seen: set = set()
    for snap in snapshots:
        for crit in snap.get("criteria_breakdown") or []:
            name = crit.get("name", "")
            if name and name not in seen:
                seen.add(name)
                criterion_names.append(name)

    base_headers = [
        "Calculated At",
        "Score Value",
        "Color",
        "Base Score",
        "Risk Penalty",
        "Risk Findings (Critical)",
        "Risk Findings (High)",
        "Risk Findings (Medium)",
        "Risk Findings (Low)",
    ]
    # Per-criterion: three columns each (points, max, achieved)
    crit_headers = []
    for name in criterion_names:
        crit_headers.extend([f"{name} — Points", f"{name} — Max", f"{name} — Achieved"])

    headers = base_headers + crit_headers
    ws.append(headers)
    _style_header_row(ws, row=1, col_count=len(headers))

    for snap in snapshots:
        rf = snap.get("risk_findings_count") or {}
        row = [
            snap.get("calculated_at", ""),
            snap.get("score_value", ""),
            snap.get("color", ""),
            snap.get("base_score", ""),
            snap.get("risk_penalty", ""),
            rf.get("critical", 0),
            rf.get("high", 0),
            rf.get("medium", 0),
            rf.get("low", 0),
        ]
        # Build a lookup so missing criteria still produce blank cells
        crit_map = {
            c.get("name", ""): c for c in (snap.get("criteria_breakdown") or [])
        }
        for name in criterion_names:
            c = crit_map.get(name, {})
            row.append(c.get("points", ""))
            row.append(c.get("max_points", ""))
            row.append("Yes" if c.get("achieved") else ("No" if "achieved" in c else ""))
        ws.append(row)

    # Auto-size columns (capped) for readability
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 40
        )

    # Freeze the header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    date_str = exported_at[:10]
    return {
        "format": "xlsx",
        "exported_at": exported_at,
        "trust_id": trust_id,
        "trust_name": trust_name,
        "xlsx_bytes": xlsx_bytes,
        "filename": f"health_scores_{trust_id}_{date_str}.xlsx",
        "snapshot_count": len(snapshots),
    }
